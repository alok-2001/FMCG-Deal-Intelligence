"""
Builds the structured newsletter as a styled .xlsx workbook -- one of the
three accepted output formats the brief asks for (excel/word/ppt).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16)
SUBTLE_FONT = Font(name="Arial", italic=True, size=9, color="6B7280")

COLUMNS = [
    ("Deal Type", "deal_type", 16),
    ("Acquirer", "acquirer", 22),
    ("Target", "target", 22),
    ("Deal Value", "deal_value", 14),
    ("Headline", "title", 55),
    ("Summary", "snippet", 70),
    ("Sources (count)", "corroboration_count", 14),
    ("Credibility Tier", "credibility_tier", 16),
    ("Credibility Score", "credibility_score", 14),
    ("Relevance Score", "relevance_score", 14),
    ("Publisher", "source", 20),
    ("Link", "url", 45),
    ("Published", "published", 14),
]


def build_excel(included_articles, path, generated_at):
    wb = Workbook()
    ws = wb.active
    ws.title = "Newsletter"

    ws.merge_cells("A1:D1")
    ws["A1"] = "FMCG Deal Intelligence Newsletter"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:D2")
    ws["A2"] = f"Generated {generated_at} — {len(included_articles)} deals"
    ws["A2"].font = SUBTLE_FONT

    header_row = 4
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    sorted_articles = sorted(
        included_articles,
        key=lambda a: (a.deal_type or "zzz", -a.relevance_score),
    )
    for row_offset, art in enumerate(sorted_articles, start=1):
        row = header_row + row_offset
        record = art.to_dict()
        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            value = record.get(field, "")
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=(field in ("title", "snippet")))

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(sorted_articles)}"

    wb.save(path)
    return path
